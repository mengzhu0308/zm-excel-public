#!/usr/bin/env python3
"""格式化 Excel 文件：字体、对齐等。

仅处理 .xlsx / .xlsm；运行 Python >= 3.9，openpyxl >= 3.1；
完整使用说明与参数定义见 SKILL.md。
"""

import argparse
import os
import sys
import traceback
from pathlib import Path

__version__ = "0.4.4"

# 默认副本后缀，集中在模块顶部定义；argparse / generate_copy_path / help 文案
# 统一引用，避免三处硬编码漂移
DEFAULT_COPY_SUFFIX = "_副本"
__openpyxl_min__ = "3.1"
__python_min__ = (3, 9)


def is_cjk_char(ch):
    """判断字符是否为 CJK（中日韩）文字或相关符号。

    控制字符（cp < 0x20）与 ASCII DEL 视为非 CJK，避免控制字符在列宽计算中被误算。
    """
    cp = ord(ch)
    if cp < 0x20 or cp == 0x7F:
        return False
    # CJK 统一表意文字
    if 0x4E00 <= cp <= 0x9FFF:
        return True
    if 0x3400 <= cp <= 0x4DBF:
        return True
    if 0x20000 <= cp <= 0x2A6DF:
        return True
    if 0x2A700 <= cp <= 0x2B73F:
        return True
    if 0x2B740 <= cp <= 0x2B81F:
        return True
    if 0x2B820 <= cp <= 0x2CEAF:
        return True
    if 0xF900 <= cp <= 0xFAFF:
        return True
    if 0x2F800 <= cp <= 0x2FA1F:
        return True
    # CJK 符号与标点
    if 0x3000 <= cp <= 0x303F:
        return True
    # 全角字符
    if 0xFF00 <= cp <= 0xFFEF:
        return True
    # 日文假名（平假名 + 片假名）
    if 0x3040 <= cp <= 0x309F:
        return True
    if 0x30A0 <= cp <= 0x30FF:
        return True
    # 韩文 Hangul
    if 0xAC00 <= cp <= 0xD7AF:
        return True
    if 0x1100 <= cp <= 0x11FF:
        return True
    if 0x3130 <= cp <= 0x318F:
        return True
    # Emoji（避免含 emoji 文本在宋体下渲染为方块）
    if 0x1F300 <= cp <= 0x1FAFF:
        return True
    if 0x2600 <= cp <= 0x27BF:
        return True
    return False


def get_font_name(text):
    """根据文本内容判断字体：含 CJK 字符用宋体，否则用 Times New Roman。"""
    if text and any(is_cjk_char(ch) for ch in text):
        return "宋体"
    return "Times New Roman"


def get_display_width(text):
    """计算文本显示宽度：CJK 等宽字符计为 2，其余可显示字符计为 1，控制字符不计。"""
    if not text:
        return 0
    width = 0
    for ch in str(text):
        cp = ord(ch)
        if cp < 0x20 or cp == 0x7F:
            continue
        if is_cjk_char(ch):
            width += 2
        else:
            width += 1
    return width


def adjust_column_widths(ws, max_width=50, min_width=8, padding=2):
    """自动调整工作表各列宽度至最优。

    按列遍历所有单元格，根据内容显示宽度（含换行处理）计算最佳列宽。
    最终列宽 = max(内容最大宽度 + padding, min_width)，但不超过 max_width。

    合并单元格处理：合并区域内的"被合并" cell（除左上角外）不参与列宽计算，
    避免合并区域跨多列时把整列拉到合并 cell 文本宽度。
    """
    from openpyxl.utils import get_column_letter

    col_max_widths = {}

    # 收集合并区域右下角（不含左上角）的 cell 坐标，用于跳过
    merged_secondary: set = set()
    for mr in ws.merged_cells.ranges:
        for row in ws[mr.coord]:
            for cell in row:
                if (cell.row, cell.column) != (mr.min_row, mr.min_col):
                    merged_secondary.add((cell.row, cell.column))

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in merged_secondary:
                continue
            col_letter = get_column_letter(cell.column)
            text = str(cell.value)
            lines = text.split("\n")
            line_width = max((get_display_width(line) for line in lines), default=0)
            current_max = col_max_widths.get(col_letter, 0)
            if line_width > current_max:
                col_max_widths[col_letter] = line_width

    for col_letter, content_width in col_max_widths.items():
        adjusted = min(content_width + padding, max_width)
        adjusted = max(adjusted, min_width)
        ws.column_dimensions[col_letter].width = adjusted


def format_workbook(input_path, output_path, adjust_width=True):
    """格式化单个 Excel 工作簿。"""
    try:
        import openpyxl
        import zipfile
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请先安装：conda install openpyxl") from e

    # 损坏 / 加密 / 格式错误的 xlsx 给用户友好提示，避免原始 openpyxl 异常
    try:
        wb = openpyxl.load_workbook(input_path)
    except zipfile.BadZipFile as e:
        raise RuntimeError(
            f"无法打开 {input_path}：文件已损坏或不是有效的 xlsx/xlsm 格式。"
            f"openpyxl 不支持 .xls 二进制格式；加密文件请先用 Excel 另存为非加密版本。"
        ) from e
    except InvalidFileException as e:
        raise RuntimeError(
            f"无法打开 {input_path}：文件可能被加密或使用了 openpyxl 不支持的格式。"
            f"如需保留宏请用 LibreOffice CLI / xlwings 等专用工具。"
        ) from e

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value) if cell.value is not None else ""
                # 单行居中，多行（含显式换行符）两端均匀对齐
                horizontal = "distributed" if "\n" in text else "center"

                # 为所有单元格设置对齐样式，避免空行/空单元格被裁剪导致行数减少
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=horizontal,
                    wrap_text=True,
                )

                if cell.value is not None:
                    font_name = get_font_name(text)
                    # 注意：cell.font.size == 0 是合法值（表示继承默认），
                    # 用 `if cell.font.size` 会把 0 误判并改成 11，因此用 `is not None` 区分
                    current_size = cell.font.size
                    new_size = current_size if current_size is not None else 11
                    cell.font = Font(
                        name=font_name,
                        size=new_size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color,
                        # 显式复制 vertAlign / scheme / family 等其他字体属性，
                        # 避免上标 / 下标 / 字体方案被重置为 None
                        vertAlign=cell.font.vertAlign,
                        scheme=cell.font.scheme,
                        family=cell.font.family,
                    )

        if adjust_width:
            adjust_column_widths(ws)

    wb.save(output_path)
    return output_path


def generate_copy_path(original, max_attempts=100, copy_suffix=DEFAULT_COPY_SUFFIX):
    """生成同目录副本路径，自动处理文件名冲突。

    - `copy_suffix` 控制副本后缀（默认 `DEFAULT_COPY_SUFFIX = "_副本"`，可通过 CLI `--copy-suffix` 改写）
    - 冲突编号 `n` 在 `1..max_attempts` 之间递增；超过则抛 RuntimeError
    - 默认 `max_attempts=100`，超出后报错并打印已尝试的候选路径
    """
    stem = original.stem
    suffix = original.suffix
    candidate = original.parent / f"{stem}{copy_suffix}{suffix}"
    if not candidate.exists():
        return candidate
    for n in range(1, max_attempts + 1):
        candidate = original.parent / f"{stem}{copy_suffix}{n}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(
        f"无法为 {original} 生成副本：同名冲突超过 {max_attempts} 次"
    )


def collect_files(paths):
    """从输入路径列表收集所有 Excel 文件。

    返回 (unique, missing_inputs)：
    - unique：去重并按白名单过滤后的 Path 列表（保留首次出现顺序）
    - missing_inputs：用户传入但找不到（既不是 file 也不是 dir，父目录
      不存在或通配符无匹配）的输入项字符串列表；调用方应将每条打印到
      stderr 提示用户

    - 空字符串输入拒绝（避免 "Path(\"\") = Path(\".\")" 被理解为"遍历当前目录"）
    - 递归通配符 `**` 显式拒绝（fnmatch 不支持 `**`；用户用 `subdir/**/*.xlsx`
      时给明确错误，避免被静默归到 missing_inputs 且错误信息不清晰）
    - 目录遍历使用 iterdir + 后缀判断，大小写不敏感（兼容 `CAPS.XLSX`）
    - 显式跳过以 `.` 开头的隐藏文件 / 隐藏目录
    - 去重并保留首次出现顺序
    """
    candidates = []
    missing_inputs = []
    for p in paths:
        # 空字符串拒绝：Path("") 等价于 Path(".")，会触发"遍历当前目录"
        if not p:
            missing_inputs.append("<空字符串>")
            continue
        # 递归通配符拒绝：fnmatch 不支持 `**`，给明确错误
        if "**" in str(p):
            missing_inputs.append(
                f"{p}（注意：递归通配符 ** 不被支持，请改用子目录作为输入）"
            )
            continue
        p_obj = Path(p)
        if p_obj.is_dir():
            for child in p_obj.iterdir():
                if child.name.startswith("."):
                    continue
                if child.is_file():
                    candidates.append(child)
        elif p_obj.is_file():
            candidates.append(p_obj)
        else:
            # 尝试通配符匹配
            parent = p_obj.parent
            matched_any = False
            if parent.exists() and parent.is_dir():
                for child in parent.iterdir():
                    if child.name.startswith("."):
                        continue
                    if child.is_file() and child.match(p_obj.name):
                        candidates.append(child)
                        matched_any = True
            if not matched_any:
                missing_inputs.append(str(p))
    # 去重（保留首次出现顺序），并按白名单过滤后缀（大小写不敏感）
    seen = set()
    unique = []
    for f in candidates:
        try:
            key = str(f.resolve())
        except OSError:
            key = str(f)
        if key in seen:
            continue
        seen.add(key)
        if f.suffix.lower() in (".xlsx", ".xlsm"):
            unique.append(f)
    return unique, missing_inputs


def compute_output_paths(files, args, copy_suffix):
    """根据 args 计算每个输入文件的输出路径。

    返回 (out_paths, errors)：
    - out_paths：与 files 等长的 Path 列表
    - errors：路径相关错误信息列表；非空时应直接退出

    行为规则：
    - `--in-place`：每个 out_path 与 f 相同
    - `--output <file>`（文件模式）：仅当 len(files) == 1 时允许；
      len(files) > 1 时报错（多文件不能共享单一文件输出）
    - `--output <dir>`（目录模式）：
      * 先按 `out / f.name` 计算候选
      * 检测候选间的冲突；冲突时把父目录名拼到 stem 后做消歧
    - 默认（无 --in-place / 无 --output）：调用 generate_copy_path(f, copy_suffix)

    路径已存在时按 is_file() 判定文件/目录模式；路径不存在时按"含后缀"判为
    文件，其余一律走目录模式（含 `./out.xlsx`、`out.xlsx`、`out.tar.gz` 等）。
    """
    errors = []
    out_paths: list = []

    if args.in_place:
        out_paths = [f for f in files]
    elif args.output:
        out_arg = Path(args.output)
        # 鲁棒文件/目录判定：路径已存在 → 看 is_file；不存在 → 含后缀视为
        # 文件（用户传 -o ./out.xlsx / out.xlsx / out.tar.gz 都算文件模式）
        if out_arg.exists():
            treat_as_file = out_arg.is_file()
        else:
            treat_as_file = bool(out_arg.suffix)
        if len(files) > 1 and treat_as_file:
            errors.append(
                f"错误：--output 指定为文件路径（{out_arg}），但输入有 {len(files)} 个文件；"
                "多文件共享单一输出文件会导致静默覆盖。请改用目录路径或单独处理每个文件。"
            )
            return [], errors

        # 单文件 + 文件路径：直接落到 out_arg
        if len(files) == 1 and treat_as_file:
            out_paths = [out_arg]
        else:
            # 目录模式：先按 f.name 算候选，冲突时拼接父目录名
            candidate_to_files = {}
            for f in files:
                cand = out_arg / f.name
                candidate_to_files.setdefault(str(cand), []).append(f)

            resolved = {}
            for cand_str, fs in candidate_to_files.items():
                cand = Path(cand_str)
                if len(fs) == 1:
                    resolved[fs[0]] = cand
                else:
                    # 多文件撞同一候选名：按父目录名消歧
                    for f in fs:
                        parent_tag = f.parent.name or "input"
                        disambig = f"{f.stem}__{parent_tag}{f.suffix}"
                        resolved[f] = out_arg / disambig

            out_paths = [resolved[f] for f in files]
    else:
        # 默认副本模式
        out_paths = [generate_copy_path(f, copy_suffix=copy_suffix) for f in files]

    # 防御性断言：消歧逻辑仍撞名（边缘 case）→ 直接报错
    if len(set(out_paths)) != len(out_paths):
        raise RuntimeError(
            f"内部错误：输出路径消歧后仍撞名：{out_paths}"
        )

    return out_paths, errors


def precheck_file(f, args):
    """对单个输入文件做资源限制预检（文件大小 + 工作表数）。

    返回 None 表示通过；返回 str 表示具体错误信息，调用方应打印到 stderr 并跳过该文件。

    设计要点：in-place 模式必须先做预检再决定是否进入 format_workbook，
    否则会出现"先原子替换原文件、再发现 sheets 超限"的数据破坏。
    """
    # 文件大小预检（不打开 xlsx 即可判断）
    try:
        if f.stat().st_size > args.max_file_size:
            return (
                f"错误：{f} 超过 --max-file-size={args.max_file_size} 字节；"
                "可调大阈值或拆文件处理。"
            )
    except OSError as e:
        return f"错误：无法读取文件信息 {f}: {e}"

    # 工作表数预检（read_only 模式打开后立即关闭，节省内存）
    try:
        import openpyxl as _ob
        _wb = _ob.load_workbook(str(f), read_only=True)
        n_sheets = len(_wb.worksheets)
        _wb.close()
        if n_sheets > args.max_sheets:
            return (
                f"错误：{f} 的工作表数（{n_sheets}）"
                f"超过 --max-sheets={args.max_sheets}。"
            )
    except Exception as e:
        return f"错误：{f} 工作表数预检失败: {e}"

    return None


def _check_runtime():
    """运行时校验 Python 与 openpyxl 版本；失败时打印明确错误并退出。

    不影响 --help（检查在 main 入口执行，argparse 解析失败会先于本检查）。
    用 tuple 元组比较避免引入 packaging 依赖。
    """
    import openpyxl

    cur_py = sys.version_info[:2]
    if cur_py < __python_min__:
        print(
            f"错误：需要 Python >={__python_min__[0]}.{__python_min__[1]}，"
            f"当前为 {cur_py[0]}.{cur_py[1]}。",
            file=sys.stderr,
        )
        sys.exit(1)
    try:
        cur_ox = tuple(int(x) for x in openpyxl.__version__.split(".")[:2])
    except (ValueError, AttributeError):
        # 解析失败时跳过 openpyxl 版本检查，避免阻塞合法环境
        return
    need_ox = tuple(int(x) for x in __openpyxl_min__.split(".")[:2])
    if cur_ox < need_ox:
        print(
            f"错误：需要 openpyxl >={__openpyxl_min__}，"
            f"当前为 {openpyxl.__version__}。",
            file=sys.stderr,
        )
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="格式化 Excel 文件：字体、对齐。"
    )
    parser.add_argument(
        "input",
        nargs="+",
        help="输入文件、目录或通配符，可指定多个",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="直接覆盖原文件（写入过程采用临时文件+原子替换，最大化降低中断风险）",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="输出目录或文件路径（与 --in-place 互斥）",
    )
    parser.add_argument(
        "--no-adjust-width",
        action="store_true",
        help="禁用自动调整列宽（默认启用）",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="显示详细日志",
    )
    parser.add_argument(
        "--copy-suffix",
        default=DEFAULT_COPY_SUFFIX,
        help="默认副本模式下的文件名后缀（默认 '_副本'）；可改为 '_copy' 等",
    )
    parser.add_argument(
        "--max-file-size",
        type=int,
        default=200 * 1024 * 1024,
        help="单文件最大字节数（默认 200MB），超出则跳过并提示",
    )
    parser.add_argument(
        "--max-sheets",
        type=int,
        default=50,
        help="工作簿最大工作表数（默认 50），超出则跳过并提示",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只打印将处理的文件与输出路径，不实际写入",
    )

    args = parser.parse_args()

    # 互斥校验（argparse 推荐：parser.error 自动打印 usage + 退出码 2）
    if args.output and args.in_place:
        parser.error("--output 与 --in-place 不能同时使用")

    # 运行时版本校验：--dry-run 模式下不触发（用户可能没有 openpyxl 也想预览路径）；
    # 写盘路径下仍校验，避免 ImportError 时已经破坏原文件
    if not args.dry_run:
        _check_runtime()

    files, missing_inputs = collect_files(args.input)

    # 输入路径错误处理：纯不存在输入时直接退出并提示；混合时打印告警但不退出
    if missing_inputs:
        for m in missing_inputs:
            print(
                f"错误：输入路径找不到（既不是文件也不是目录，父目录也不存在或通配符无匹配）：{m}",
                file=sys.stderr,
            )
        if not files:
            print("错误：所有输入路径都未找到，未处理任何文件。", file=sys.stderr)
            sys.exit(1)
        print(
            f"提示：以上 {len(missing_inputs)} 个输入被忽略，其余 {len(files)} 个文件继续处理。",
            file=sys.stderr,
        )

    if not files:
        print("错误：未找到任何 Excel 文件。", file=sys.stderr)
        sys.exit(1)

    out_paths, path_errors = compute_output_paths(files, args, args.copy_suffix)
    for err in path_errors:
        print(err, file=sys.stderr)
    if path_errors:
        sys.exit(1)

    # 资源限制预检：循环外统一做，避免 in-place 模式先破坏原文件再发现超限。
    # --dry-run 模式下不预检（避免对每个文件 load_workbook(read_only=True) 触发
    # 的 N 次 I/O；dry-run 用户预期是秒级预览）
    failed_count = 0
    if not args.dry_run:
        total = len(files)
        filtered_files = []
        filtered_out_paths = []
        for f, op in zip(files, out_paths):
            err = precheck_file(f, args)
            if err is not None:
                print(err, file=sys.stderr)
                failed_count += 1
                continue
            filtered_files.append(f)
            filtered_out_paths.append(op)
        if not filtered_files:
            print("错误：所有输入文件都因资源限制被跳过，未处理任何文件。", file=sys.stderr)
            sys.exit(1)
        files = filtered_files
        out_paths = filtered_out_paths
        total = len(files)
    else:
        total = len(files)

    for idx, (f, out_path) in enumerate(zip(files, out_paths), 1):
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if args.dry_run:
            print(f"[dry-run] {idx}/{total} {f} -> {out_path}")
            continue

        # 所有非 dry-run 路径都走 atomic 包装：先写临时文件再原子替换，
        # 避免极端中断（磁盘满 / 编码异常 / 进程被杀）留下半成品
        in_place = args.in_place and out_path == f
        tmp_path = (
            f.with_name(f".{f.name}.{os.getpid()}.{idx}.tmp")
            if in_place
            else out_path.with_name(f".{out_path.name}.{os.getpid()}.{idx}.tmp")
        )
        try:
            format_workbook(
                str(f), str(tmp_path), adjust_width=not args.no_adjust_width
            )
            os.replace(tmp_path, out_path)
            if args.verbose:
                tag = "（原子替换）" if in_place else ""
                target = f if in_place else out_path
                print(f"已格式化{tag}: {f} -> {target}")
        except Exception as e:
            print(f"错误：处理 {f} 失败: {e}", file=sys.stderr)
            if args.verbose:
                traceback.print_exc()
            failed_count += 1
            if total == 1:
                sys.exit(1)
            continue
        finally:
            # finally 在 Exception / KeyboardInterrupt / SystemExit 下都会执行
            # 防止 Ctrl+C 时留下孤儿临时文件；此处统一处理，except 段不再重复
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass

        if args.verbose and total > 1:
            print(f"[{idx}/{total}] 已完成: {f.name}")

    # 多文件部分失败时退出码非零（CI/CD 可据此判定"全部成功 vs 部分失败"）
    if failed_count > 0:
        print(
            f"完成：共处理 {total} 个文件，其中 {failed_count} 个失败。",
            file=sys.stderr,
        )
        sys.exit(1)
    print(f"完成：共处理 {total} 个文件。")


if __name__ == "__main__":
    main()

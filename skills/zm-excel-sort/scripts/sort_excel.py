#!/usr/bin/env python3
"""
zm-excel-sort 执行脚本
支持按字段规则对 Excel/CSV 进行排序，XLSX 输出保留原始样式。
"""

import argparse
import json
import sys
import traceback
from copy import copy
from pathlib import Path


def log(msg, verbose=False):
    if verbose:
        print(f"[zm-excel-sort] {msg}", file=sys.stderr)


def _escape_csv_formula(values):
    """CSV 注入防护：对以危险前缀开头的字段加单引号前缀

    Excel/WPS 在打开 CSV 时会把 `=` / `+` / `-` / `@` 开头的字段当公式执行，
    可能改变原始数据值或触发恶意公式。对这些字段加 `'` 前缀，
    强制 Excel 视为普通文本。
    """
    dangerous = ("=", "+", "-", "@", "\t", "\r")
    escaped = []
    for v in values:
        if isinstance(v, str) and v[:1] in dangerous:
            escaped.append("'" + v)
        else:
            escaped.append(v)
    return escaped


def _ensure_parent_dir(output_path):
    """若输出路径的父目录不存在则自动创建，避免依赖底层库抛 FileNotFoundError"""
    parent = Path(output_path).expanduser().resolve().parent
    if not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def parse_args():
    parser = argparse.ArgumentParser(description="Sort Excel/CSV by field rules")
    parser.add_argument("--input", "-i", required=True, help="输入文件路径")
    parser.add_argument("--output", "-o", required=True, help="输出文件路径")
    parser.add_argument("--rules", "-r", required=True, help="排序规则 JSON")
    parser.add_argument("--sheet", "-s", default="0", help="Sheet 名称或索引")
    parser.add_argument("--format", "-f", choices=["csv", "xlsx"], help="输出格式")
    parser.add_argument("--verbose", "-v", action="store_true", help="详细日志")
    return parser.parse_args()


def detect_format(filepath, explicit_format=None):
    if explicit_format:
        return explicit_format
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        return "csv"
    elif ext in (".xlsx", ".xlsm"):
        # openpyxl 原生支持 .xlsx 与 .xlsm（带宏）；skill 与 SKILL.md/README.md 承诺一致
        return "xlsx"
    elif ext == ".xls":
        # openpyxl 不支持 .xls：直接拒绝，避免在 openpyxl.load_workbook 阶段抛
        raise ValueError(
            f"不支持的 Excel 格式 '{ext}'。本 skill 仅支持 .xlsx（及 .xlsm）。"
            f"如需处理 .xls，请先用 Excel/WPS/LibreOffice 另存为 .xlsx 后再调用。"
        )
    else:
        raise ValueError(
            f"无法从扩展名推断格式: '{ext}'。仅支持 .csv、.xlsx、.xlsm。"
        )


def load_input(filepath, sheet_spec, verbose=False):
    """加载输入文件，返回 (headers, data_rows, meta)"""
    fmt = detect_format(filepath)

    if fmt == "csv":
        # P0-1：CSV 不支持 --sheet；非默认 "0" 直接拒绝，避免静默忽略
        if sheet_spec != "0":
            raise ValueError(
                f"--sheet '{sheet_spec}' 仅对 Excel 输入有效，当前输入为 CSV。"
                "CSV 文件没有 Sheet 概念，请去掉 --sheet 参数。"
            )
        import pandas as pd
        # 尝试 UTF-8 带 BOM，然后普通 UTF-8，最后回退常见中文编码
        encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030", "latin1"]
        df = None
        enc = encodings[0]  # 默认初始化，避免 enc unbound 警告（真正使用的值由 for 循环赋值）
        for enc in encodings:
            try:
                df = pd.read_csv(filepath, encoding=enc)
                # P1-7：正常编码走 verbose；仅 latin1 兜底时强制 stderr 提示，避免静默吞错
                if enc == "latin1":
                    print(
                        f"[zm-excel-sort] CSV 编码检测: {enc}（兜底；可能存在乱码，"
                        f"建议显式指定 UTF-8 / GBK / GB18030 等真实编码）",
                        file=sys.stderr,
                    )
                else:
                    log(f"CSV 使用编码: {enc}", verbose)
                break
            except UnicodeDecodeError:
                continue
        # 注：原循环之后曾含一个"df is None 时 raise ValueError"的死代码分支。
        # A-2 P0-2 修复：删除该分支——latin1 是 0-255 全字符编码，任意字节序列都能解码，
        # 该错误永远不会被触发；仅保留此处注释作为变更轨迹。同时把 latin1 兜底提示
        # 改为更明确的乱码警告。

        headers = list(df.columns)
        data_rows = df.values.tolist()
        meta = {"source": "csv", "encoding": enc}
        return headers, data_rows, meta

    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)

        # 解析 sheet 指定：先按名匹配，匹配失败后才回退到数字索引。
        # 这样字面 Sheet 名 "1" 不会再被 `isdigit()` 吞掉解释为索引 1。
        if sheet_spec in wb.sheetnames:
            ws_name = sheet_spec
        elif sheet_spec.isdigit():
            idx = int(sheet_spec)
            if idx < 0 or idx >= len(wb.sheetnames):
                available = ", ".join(wb.sheetnames)
                raise ValueError(f"Sheet 索引 {idx} 超出范围。可用 Sheet: {available}")
            ws_name = wb.sheetnames[idx]
        else:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"Sheet '{sheet_spec}' 不存在。可用 Sheet: {available}")

        ws = wb[ws_name]
        log(f"使用 Sheet: {ws_name}", verbose)

        # 提取样式信息（P0-5 修复：先初始化 styles，再写入 merged_cells，避免 unbound）
        # P1-2 修复：data_validations copy 改为 try/except + log 模式，与 A-1 P0-3 修复 tables 同风格；
        # 老 openpyxl 中 dv.copy() 在 DataValidation 含 formula1 复杂表达式时可能抛 AttributeError，
        # 整条 load_input 链崩溃——改为单条失败单条 log，不阻断 load_input
        dv_copies = []
        for _dv in ws.data_validations.dataValidation:
            try:
                dv_copies.append(copy(_dv))
            except Exception as e:
                log(f"复制 DataValidation '{_dv}' 失败（{type(e).__name__}: {e}）；跳过", verbose)
        styles = {
            "column_dimensions": {},
            "row_dimensions": {},
            "cell_styles": [],
            "cell_comments": [],
            "cell_hyperlinks": [],
            # 排序后会丢失或失真的项：先抽取再在 write_xlsx 中尝试保留
            "freeze_panes": ws.freeze_panes,
            "conditional_formatting": [],
            "data_validations": dv_copies,
            "tables": [],
            "merged_cells": [],
        }

        # P0-5 修复：ws.merged_cells.ranges 在 openpyxl < 3.0.5 不存在
        # 统一在 styles 字典中记录，write_xlsx 阶段再按 ranges/字符串协议使用
        try:
            merged_ranges = list(ws.merged_cells.ranges)
        except AttributeError:
            # 老 API：ws.merged_cells 自身是 MergedCellRange 集合
            try:
                merged_ranges = list(ws.merged_cells)
            except Exception as e:
                log(f"读取合并单元格失败（{type(e).__name__}: {e}）；将不复制合并", verbose)
                merged_ranges = []
        styles["merged_cells"] = merged_ranges

        # 读取所有数据
        all_rows = list(ws.iter_rows(values_only=False))
        if not all_rows:
            return [], [], {"source": "xlsx", "workbook": wb, "worksheet": ws, "sheet_name": ws_name}

        headers = [cell.value for cell in all_rows[0]]
        # P1-3 修复：合并表头场景下，MergedCell（非左上角）取值返回 None
        # 保留 None 占位以保持列数对齐；DataFrame 构造时 openpyxl / pandas 会用 NaN 占位
        # 后续如果 column 名称为 None，sort_data 会让 AI 据此回退到用户确认（已由字段不存在校验覆盖）
        data_rows = all_rows[1:]

        # P1-4：访问 openpyxl 私有 API `_cf_rules` 加 try/except + log 兜底
        try:
            cf_rules = ws.conditional_formatting._cf_rules
            # B-1 修复：_cf_rules 的 key 是 ConditionalFormatting 包装类；
            # key.sqref 是 MultiCellRange，str() 后拿干净 'A2:A3' 字符串；
            # add() 接 str 时会包装为 ConditionalFormatting，存盘走 to_tree() 正常；
            # 不能传 MultiCellRange（add() 不会包装，存盘 to_tree() 崩 'MultiCellRange' has no 'to_tree'）
            styles["conditional_formatting"] = [
                (str(key.sqref), [copy(rule) for rule in rule_list])
                for key, rule_list in cf_rules.items()
            ]
        except Exception as e:
            log(f"读取条件格式失败（{type(e).__name__}: {e}）；将不复制条件格式", verbose)

        # P0-3：含 Table 的 XLSX 触发 Table.__copy__ TypeError；改用 try/except + log
        # 策略：保留 Table 的字符串表示（displayName + ref）以便 write_xlsx 重新 add_table；
        # 若 __copy__ 失败则跳过并 log
        from openpyxl.worksheet.table import Table
        for tbl in ws.tables.values():
            try:
                styles["tables"].append(copy(tbl))
            except Exception as e:
                log(f"复制 Table '{tbl}' 失败（{type(e).__name__}: {e}）；跳过", verbose)

        for col_letter in ws.column_dimensions:
            dim = ws.column_dimensions[col_letter]
            styles["column_dimensions"][col_letter] = {
                "width": dim.width,
                "hidden": dim.hidden,
            }

        for row_idx, row in enumerate(all_rows, start=1):
            dim = ws.row_dimensions[row_idx]
            styles["row_dimensions"][row_idx] = {
                "height": dim.height,
                "hidden": dim.hidden,
            }
            row_styles = []
            for cell in row:
                row_styles.append({
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                })
                # P2-4：保留 cell.comment 与 cell.hyperlink
                if cell.comment is not None:
                    styles["cell_comments"].append(
                        (row_idx, cell.column, copy(cell.comment))
                    )
                if cell.hyperlink is not None:
                    styles["cell_hyperlinks"].append(
                        (row_idx, cell.column, copy(cell.hyperlink))
                    )
            styles["cell_styles"].append(row_styles)

        meta = {
            "source": "xlsx",
            "workbook": wb,
            "worksheet": ws,
            "sheet_name": ws_name,
            "styles": styles,
        }
        return headers, data_rows, meta


def sort_data(headers, data_rows, rules, verbose=False):
    """根据规则排序数据，返回排序后的索引列表"""
    import pandas as pd

    if not data_rows:
        return []

    # P1-NEW-4：表头含重复列名时直接拒绝
    # A-2 修复：把 None/NaN 排除后再做重复检查——合并表头场景下 P1-3 修复会保留 None 占位，
    # pandas 会把 None 列名转成 NaN，None/NaN 在 headers 中出现多次是正常的，不应被误判为重复列名。
    non_null_headers = [h for h in headers if h is not None and not (isinstance(h, float) and pd.isna(h))]
    if len(non_null_headers) != len(set(non_null_headers)):
        from collections import Counter
        dup = [v for v, c in Counter(non_null_headers).items() if c > 1]
        raise ValueError(
            f"输入表头含重复列名: {dup}。请先在源文件中重命名重复列，再调用本 skill。"
        )

    # 创建 DataFrame
    df = pd.DataFrame(
        [[cell.value if hasattr(cell, "value") else cell for cell in row] for row in data_rows],
        columns=headers,
    )

    # P1-NEW-3：检测每列数据类型混合（如文本+数字），统一按文本排序并提示
    # SKILL.md 错误处理表格第 7 行承诺"统一按文本排序，并提示用户"——
    # 提示应不依赖 --verbose（用户不开 verbose 也要看到），故走 stderr
    # A-2 修复：排除 None/NaN 列——合并表头场景下 P1-3 修复保留 None 占位，
    # pandas 会把 None 列名转成 NaN，df[NaN] 在多个 NaN 列时返回 DataFrame，访问 .dtype 会 AttributeError。
    for col in df.columns:
        if col is None or (isinstance(col, float) and pd.isna(col)):
            continue
        if df[col].dtype == object:
            non_null = df[col].dropna()
            if non_null.empty:
                continue
            types = {type(v).__name__ for v in non_null}
            if len(types) > 1:
                print(
                    f"[zm-excel-sort] 提示: 列 '{col}' 含多种 Python 类型 {sorted(types)}，"
                    f"将统一按文本排序。建议在源文件中统一为单一类型再调用。",
                    file=sys.stderr,
                )

    # === 规则 schema 校验（P1-8 / P1-9） ===
    if not isinstance(rules, dict):
        raise ValueError("排序规则必须是 JSON 对象，例如 {\"columns\": [...]}")

    if "columns" not in rules:
        raise ValueError(
            "排序规则缺少 'columns' 字段；至少应包含 1 个排序字段，例如 "
            "{\"columns\":[{\"name\":\"销售额\",\"direction\":\"desc\"}]}"
        )
    columns = rules["columns"]
    if not isinstance(columns, list) or len(columns) == 0:
        # A-2 修复：排除 NaN/None 列名后 join——合并表头场景下 df.columns 含 NaN float，
        # ", ".join 时会 TypeError: sequence item 1: expected str instance, float found
        safe_cols = [str(c) for c in df.columns if not (c is None or (isinstance(c, float) and pd.isna(c)))]
        available = ", ".join(safe_cols)
        raise ValueError(
            f"'columns' 必须是非空列表；当前为空，将等同于不排序。可用字段: {available}"
        )

    null_pos = rules.get("null_position", "last")
    if null_pos not in ("first", "last"):
        raise ValueError(
            f"'null_position' 仅支持 'first' 或 'last'，当前为 '{null_pos}'"
        )
    na_position = "first" if null_pos == "first" else "last"

    case_sensitive = rules.get("case_sensitive", False)
    if not isinstance(case_sensitive, bool):
        raise ValueError(f"'case_sensitive' 必须是布尔值，当前为 {case_sensitive!r}")

    # === 字段级校验 ===
    sort_columns = []
    ascending_flags = []
    custom_order_columns = set()  # 记录已设置 custom_order 的列，避免大小写处理冲突

    for col_rule in columns:
        if not isinstance(col_rule, dict):
            raise ValueError(
                f"排序规则项必须是对象，例如 {{\"name\":\"X\",\"direction\":\"asc\"}}；当前: {col_rule!r}"
            )
        if "name" not in col_rule:
            raise ValueError(f"排序规则项缺少 'name' 字段: {col_rule!r}")

        col_name = col_rule["name"]
        if col_name not in df.columns:
            # A-2 修复：排除 NaN/None 列名后 join
            safe_cols = [str(c) for c in df.columns if not (c is None or (isinstance(c, float) and pd.isna(c)))]
            available = ", ".join(safe_cols)
            raise ValueError(f"排序字段 '{col_name}' 不存在。可用字段: {available}")

        direction = col_rule.get("direction", "asc")
        if not isinstance(direction, str) or direction.lower() not in ("asc", "desc"):
            raise ValueError(
                f"排序字段 '{col_name}' 的 direction 仅支持 'asc' 或 'desc'，当前为 {direction!r}"
            )
        ascending = direction.lower() == "asc"

        custom_order = col_rule.get("custom_order")
        if custom_order is not None:
            if not isinstance(custom_order, list) or len(custom_order) == 0:
                raise ValueError(
                    f"排序字段 '{col_name}' 的 custom_order 必须是非空列表"
                )
            # 重复项校验：pd.Categorical 不允许 categories 含重复
            if len(custom_order) != len(set(custom_order)):
                from collections import Counter
                dup = [v for v, c in Counter(custom_order).items() if c > 1]
                raise ValueError(
                    f"排序字段 '{col_name}' 的 custom_order 含重复值: {dup}"
                )
            # P1-4 修复：custom_order 元素含 None 提前拒绝
            # pd.Categorical 在 categories 含 None 时行为未定义；co_non_str 分支的 pd.to_numeric(None)
            # 会抛 TypeError；统一提前 ValueError 提示用户
            if any(v is None for v in custom_order):
                raise ValueError(
                    f"排序字段 '{col_name}' 的 custom_order 含 None；请用字符串或显式值替代 None。"
                )
            # P1-2 修复：源列与 custom_order 类型不匹配时（如 int 列 + str custom_order），
            # pd.Categorical 会把所有源值变 NaN，排序静默退化为原顺序。
            # 防御：将源列非空值归一化为与 custom_order 元素一致的字符串形式（避免 1 vs "1" 失配）。
            # P1-3 修复：custom_order 同时含 str 与非 str 时（如 ["低", 1, "高"]），单边归一化失败；
            # 改为"先看混合，再走主导类型"
            sample = df[col_name].dropna()
            if not sample.empty:
                co_non_str = [v for v in custom_order if not isinstance(v, str)]
                co_str = [v for v in custom_order if isinstance(v, str)]
                if co_non_str and co_str:
                    # P1-3：custom_order 同时含 str 和非 str；统一把所有元素与源列转 str
                    custom_order = [str(v) for v in custom_order]
                    if sample.dtype != object:
                        df[col_name] = df[col_name].astype(str)
                elif co_non_str:
                    # custom_order 全是非 str（如 int/float），把源列也统一到非 str
                    if sample.dtype == object:
                        df[col_name] = pd.to_numeric(df[col_name], errors="coerce")
                else:
                    # custom_order 全是 str；源列若不是 object（如 int/float/datetime），归一为 str
                    if sample.dtype != object:
                        df[col_name] = df[col_name].astype(str)
            # 自定义排序：将列转换为 Categorical 类型
            df[col_name] = pd.Categorical(
                df[col_name], categories=custom_order, ordered=True
            )
            custom_order_columns.add(col_name)

        sort_columns.append(col_name)
        ascending_flags.append(ascending)

    # === 排序执行 ===
    if sort_columns:
        if not case_sensitive:
            actual_sort_cols = []
            for col in sort_columns:
                # 自定义排序列（Categorical）不能也不需要再做大小写归一化
                if col in custom_order_columns:
                    actual_sort_cols.append(col)
                    continue
                # A-2 修复：排除 None/NaN 列——合并表头场景下 df[None] 在多个 None 列时返回 DataFrame，
                # 访问 .dtype 会 AttributeError
                if col is None or (isinstance(col, float) and pd.isna(col)):
                    actual_sort_cols.append(col)
                    continue
                if df[col].dtype == object:
                    # 文本列：生成小写副本用于大小写不敏感排序
                    try:
                        sort_col = f"__sort_{col}"
                        df[sort_col] = df[col].astype(str).str.lower()
                        actual_sort_cols.append(sort_col)
                    except (ValueError, TypeError, AttributeError) as e:
                        log(
                            f"列 '{col}' 跳过大小写归一化（{type(e).__name__}: {e}）",
                            verbose,
                        )
                        actual_sort_cols.append(col)
                else:
                    actual_sort_cols.append(col)

            df_sorted = df.sort_values(
                by=actual_sort_cols,
                ascending=ascending_flags,
                na_position=na_position,
            )
            # 删除辅助排序列
            for col in sort_columns:
                sort_col = f"__sort_{col}"
                if sort_col in df_sorted.columns:
                    df_sorted = df_sorted.drop(columns=[sort_col])
        else:
            df_sorted = df.sort_values(
                by=sort_columns,
                ascending=ascending_flags,
                na_position=na_position,
            )
    else:
        df_sorted = df

    # 返回排序后的索引
    return df_sorted.index.tolist()


def write_csv(headers, data_rows, sorted_indices, output_path, verbose=False):
    """写入 CSV 文件，UTF-8 with BOM"""
    import csv

    _ensure_parent_dir(output_path)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(_escape_csv_formula(headers))
        for idx in sorted_indices:
            row = data_rows[idx]
            values = [cell.value if hasattr(cell, "value") else cell for cell in row]
            writer.writerow(_escape_csv_formula(values))

    log(f"CSV 输出完成: {output_path}", verbose)


def write_xlsx(headers, data_rows, sorted_indices, meta, output_path, verbose=False):
    """写入 XLSX 文件，保留原始样式"""
    import openpyxl

    _ensure_parent_dir(output_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta.get("sheet_name", "Sheet1")

    old_styles = meta.get("styles", {})
    old_cell_styles = old_styles.get("cell_styles", [])
    old_col_dims = old_styles.get("column_dimensions", {})
    old_row_dims = old_styles.get("row_dimensions", {})
    old_merged = old_styles.get("merged_cells", [])

    # 准备所有行（包含表头）
    all_old_rows = []
    if old_cell_styles:
        all_old_rows.append(old_cell_styles[0])  # 表头行
        for idx in sorted_indices:
            all_old_rows.append(old_cell_styles[idx + 1])  # 数据行（索引+1因为第0行是表头）

    # P2-4：按 (new_row, col) 索引 cell.comment / cell.hyperlink，写完后回填
    # P0-3 修复：先初始化字典，让表头循环也能写入（第 1 行参与收集）
    new_cell_comments = {}
    new_cell_hyperlinks = {}

    # 写入表头
    # P0-3 修复：表头行也参与 comment / hyperlink 收集（之前只覆盖数据行，导致表头批注与超链接丢失）
    # P1-1 修复：表头批注 / 超链接按 (col) 建索引，循环内 O(1) 查找，避免 O(header_count × total) 重复扫描
    header_comments_by_col = {
        c: payload for (r, c, payload) in old_styles.get("cell_comments", []) if r == 1
    }
    header_hyperlinks_by_col = {
        c: payload for (r, c, payload) in old_styles.get("cell_hyperlinks", []) if r == 1
    }
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if old_cell_styles:
            style = old_cell_styles[0][col_idx - 1]
            cell.font = style["font"]
            cell.fill = style["fill"]
            cell.border = style["border"]
            cell.alignment = style["alignment"]
            cell.number_format = style["number_format"]
            cell.protection = style["protection"]
        # P0-3：表头批注与超链接在 load_input 中已按 (row_idx=1, col, ...) 存入 styles.cell_comments / hyperlinks
        # 表头永远在第 1 行；P1-1 修复后用预建索引 O(1) 取
        if col_idx in header_comments_by_col:
            new_cell_comments[(1, col_idx)] = copy(header_comments_by_col[col_idx])
        if col_idx in header_hyperlinks_by_col:
            new_cell_hyperlinks[(1, col_idx)] = copy(header_hyperlinks_by_col[col_idx])

    # 写入数据行（按排序后顺序）
    for new_row_idx, old_data_idx in enumerate(sorted_indices, start=2):
        row = data_rows[old_data_idx]
        for col_idx, cell_val in enumerate(row, start=1):
            value = cell_val.value if hasattr(cell_val, "value") else cell_val
            cell = ws.cell(row=new_row_idx, column=col_idx, value=value)
            if old_cell_styles:
                style = old_cell_styles[old_data_idx + 1][col_idx - 1]
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.border = style["border"]
                cell.alignment = style["alignment"]
                cell.number_format = style["number_format"]
                cell.protection = style["protection"]
            # P2-4：收集原 cell 的 comment / hyperlink 到新位置
            if hasattr(cell_val, "comment") and cell_val.comment is not None:
                new_cell_comments[(new_row_idx, col_idx)] = copy(cell_val.comment)
            if hasattr(cell_val, "hyperlink") and cell_val.hyperlink is not None:
                new_cell_hyperlinks[(new_row_idx, col_idx)] = copy(cell_val.hyperlink)

    # P2-4：回填 comment / hyperlink（按新行新列坐标）
    for (r, c), cm in new_cell_comments.items():
        try:
            ws.cell(row=r, column=c).comment = cm
        except Exception as e:
            log(f"comment 写入 ({r},{c}) 失败（{type(e).__name__}: {e}）", verbose)
    for (r, c), hl in new_cell_hyperlinks.items():
        try:
            ws.cell(row=r, column=c).hyperlink = hl
        except Exception as e:
            log(f"hyperlink 写入 ({r},{c}) 失败（{type(e).__name__}: {e}）", verbose)

    # 复制列宽
    for col_letter, dim_info in old_col_dims.items():
        ws.column_dimensions[col_letter].width = dim_info.get("width")
        ws.column_dimensions[col_letter].hidden = dim_info.get("hidden", False)

    # 复制行高（表头 + 数据行）
    if old_row_dims:
        # 表头行高
        if 1 in old_row_dims:
            ws.row_dimensions[1].height = old_row_dims[1].get("height")
            ws.row_dimensions[1].hidden = old_row_dims[1].get("hidden", False)
        # 数据行高（使用排序后对应的旧行高）
        for new_row_idx, old_data_idx in enumerate(sorted_indices, start=2):
            old_row_num = old_data_idx + 2  # 原始行号 = 数据索引 + 2（第1行是表头）
            if old_row_num in old_row_dims:
                ws.row_dimensions[new_row_idx].height = old_row_dims[old_row_num].get("height")
                ws.row_dimensions[new_row_idx].hidden = old_row_dims[old_row_num].get("hidden", False)

    # 处理合并单元格
    # 优先按名/索引查找 old_data → new_data 映射；多行合并在排序后保持连续时整体平移
    old_to_new = {old: new for new, old in enumerate(sorted_indices, start=2)}

    for merged_range in old_merged:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row == 1 and max_row == 1:
            # 只涉及表头，保持不变
            ws.merge_cells(str(merged_range))
            # P2-1 修复：成功路径加 verbose log（之前仅失败路径 log）
            log(f"合并区域 {merged_range} 仅涉及表头，已保留", verbose)
        elif min_row == 1 and max_row > 1:
            # 跨表头+首行合并（如 A1:B2）：表头在第 1 行，"延伸"行（max_row）是首个数据行。
            # 排序后整组数据连续平移，所以合并区应覆盖"表头 + 全部数据行"。
            # A-2 修复：之前用 old_to_new[max_row-2] 作 new_end，丢失后续数据延伸。
            new_start = 1
            new_end = 1 + len(sorted_indices)
            # new_end >= max(max_row, 2) 时才有意义（至少延伸到原 max_row）
            if new_end >= max(max_row, 2):
                ws.merge_cells(
                    start_row=1, start_column=min_col,
                    end_row=new_end, end_column=max_col,
                )
                log(f"合并区域 {merged_range} 已平移至 1..{new_end}", verbose)
            else:
                log(f"合并区域 {merged_range} 排序后无法定位首行，跳过", verbose)
        elif min_row > 1:
            old_data_start = min_row - 2
            old_data_end = max_row - 2
            if old_data_start not in old_to_new or old_data_end not in old_to_new:
                log(f"合并区域 {merged_range} 涉及越界数据行，跳过", verbose)
                continue
            new_start = old_to_new[old_data_start]
            new_end = old_to_new[old_data_end]
            # 仅当排序后该组行仍连续时整体平移；否则跳过
            if new_end - new_start == max_row - min_row:
                ws.merge_cells(
                    start_row=new_start, start_column=min_col,
                    end_row=new_end, end_column=max_col,
                )
                log(f"合并区域 {merged_range} 已平移至 {new_start}..{new_end}", verbose)
            else:
                log(
                    f"合并区域 {merged_range} 在排序后不再连续（new {new_start}..{new_end}）"
                    f"，跳过以避免语义错位",
                    verbose,
                )

    # 复制冻结窗口（P1-3）
    freeze = old_styles.get("freeze_panes")
    if freeze:
        ws.freeze_panes = freeze

    # 复制条件格式——P1-3 + P1-NEW-5：基于固定区域，排序后可能错位；显式 log 风险
    cf_items = old_styles.get("conditional_formatting", [])
    if cf_items:
        log(
            f"复制 {len(cf_items)} 个条件格式；存在错位风险——"
            f"条件格式基于固定区域引用，排序后该规则可能不再对应原行，请人工评估是否需要重新设定。",
            verbose,
        )
    for cf_range, cf_rules in cf_items:
        # B-1 修复：ws.conditional_formatting.add(range, cfRule) 接受单数 Rule，不是 list
        # 之前传整个 list 触发 ValueError "Only instances of openpyxl.formatting.rule.Rule may be added"
        # B-3 升级：失败时除 log 外还走 stderr 警告，与 DV / 命名区域 写入失败反馈对齐，
        # 让普通 verbose 模式用户也能感知"条件格式可能不完整"
        for cf_rule in cf_rules:
            try:
                ws.conditional_formatting.add(cf_range, cf_rule)
            except Exception as e:
                log(f"条件格式 '{cf_range}' 复制失败（{type(e).__name__}: {e}）", verbose)
                print(
                    f"警告: 条件格式 '{cf_range}' 复制失败（{type(e).__name__}: {e}）；"
                    f"输出 XLSX 的条件格式可能不完整，请人工评估是否需要重新设定。",
                    file=sys.stderr,
                )

    # 复制数据验证（P1-3）
    # P1-3 修复：失败时除 log 外还走 stderr 警告，避免用户感知不到 SKILL.md L117 "保留数据验证" 承诺的失真
    for dv in old_styles.get("data_validations", []):
        try:
            ws.add_data_validation(dv)
        except Exception as e:
            log(f"数据验证 '{dv}' 复制失败（{type(e).__name__}: {e}）", verbose)
            print(
                f"警告: 数据验证复制失败（{type(e).__name__}: {e}）；"
                f"输出 XLSX 的数据验证可能不完整，请人工评估是否需要重新设定。",
                file=sys.stderr,
            )

    # 复制表格（P1-3）
    for tbl in old_styles.get("tables", []):
        try:
            ws.add_table(tbl)
        except Exception as e:
            log(f"表格 '{tbl}' 复制失败（{type(e).__name__}: {e}）", verbose)

    # 复制工作簿级 defined names（命名区域）（P1-3）
    # P0-4 修复：openpyxl 3.1+ 的 DefinedNameDict 不再支持直接 __setitem__，需用 .add() / .append()
    # 此处统一走 .add()；老版本无 .add() 时回退到 __setitem__；双重 try/except 兜底
    # P1-4 修复：失败时除 log 外还走 stderr 警告，main() 拿不到失败信号的旧问题
    src_wb = meta.get("workbook")
    if src_wb is not None:
        for name, defn in src_wb.defined_names.items():
            copied = False
            try:
                wb.defined_names.add(defn)
                copied = True
            except (AttributeError, TypeError):
                # 老 API：defined_names 是 dict-like，可直接 __setitem__
                try:
                    wb.defined_names[name] = defn
                    copied = True
                except Exception as e2:
                    log(f"命名区域 '{name}' 回退复制失败（{type(e2).__name__}: {e2}）", verbose)
                    print(
                        f"警告: 命名区域 '{name}' 复制失败（{type(e2).__name__}: {e2}）；"
                        f"输出 XLSX 中引用此命名区域的公式可能 #REF!，请人工评估。",
                        file=sys.stderr,
                    )
            except Exception as e:
                log(f"命名区域 '{name}' 复制失败（{type(e).__name__}: {e}）", verbose)
                print(
                    f"警告: 命名区域 '{name}' 复制失败（{type(e).__name__}: {e}）；"
                    f"输出 XLSX 中引用此命名区域的公式可能 #REF!，请人工评估。",
                    file=sys.stderr,
                )
            if not copied:
                log(f"命名区域 '{name}' 未能复制（API 兼容性原因）", verbose)
                print(
                    f"警告: 命名区域 '{name}' 未能复制（API 兼容性原因）；"
                    f"输出 XLSX 中引用此命名区域的公式可能 #REF!，请人工评估。",
                    file=sys.stderr,
                )

    wb.save(output_path)
    log(f"XLSX 输出完成: {output_path}", verbose)


def main():
    args = parse_args()
    verbose = args.verbose

    log(f"输入: {args.input}", verbose)
    log(f"输出: {args.output}", verbose)

    # P0-NEW-2 + P1-5：--format 显式传入时，与 --output 扩展名必须一致；否则报错。
    # P1-5 修复：detect_format 三处重复调用（L625 之后 + L660 空文件路径 + L689 写入路径）合并为单一 out_fmt 变量
    out_ext = Path(args.output).suffix.lower()
    if args.format is not None:
        expected_ext = f".{args.format}"
        # .xlsm 在 detect_format 中归并为 xlsx，但用户传 --format 时仍按字面 .xlsm 校验
        if out_ext not in (expected_ext, ".xlsm" if args.format == "xlsx" else ""):
            print(
                f"错误: --format '{args.format}' 与 --output 扩展名 '{out_ext}' 不一致。"
                f"请明确选择：--format csv 时输出应为 .csv；--format xlsx 时输出应为 .xlsx 或 .xlsm。"
                f"如确需覆盖，请去掉 --format 参数，让脚本从扩展名自动推断。",
                file=sys.stderr,
            )
            sys.exit(1)

    # P1-5：提前解析输出格式供后续三处复用，避免 detect_format 三次调用
    out_fmt = detect_format(args.output, args.format)

    # 解析排序规则
    try:
        rules = json.loads(args.rules)
    except json.JSONDecodeError as e:
        print(f"错误: 排序规则 JSON 解析失败 - {e}", file=sys.stderr)
        sys.exit(1)

    log(f"排序规则: {json.dumps(rules, ensure_ascii=False)}", verbose)

    # 加载输入
    try:
        headers, data_rows, meta = load_input(args.input, args.sheet, verbose)
    except Exception as e:
        # P1-4 修复：保留 traceback，方便 AI / 用户排查
        traceback.print_exc(file=sys.stderr)
        print(f"错误: 加载输入文件失败 - {e}", file=sys.stderr)
        sys.exit(1)

    if not headers:
        print("警告: 输入文件为空，输出空文件", file=sys.stderr)
        # 写入空文件
        if out_fmt == "csv":
            write_csv([], [], [], args.output, verbose)
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            # P1-6：不传 --sheet（默认 "0"）时，输出 XLSX 的 Sheet 名兜底为 "Sheet1"；
            # 用户显式传 --sheet 时按字面值
            sheet_name = "Sheet1" if args.sheet == "0" else args.sheet
            wb.active.title = sheet_name
            log(f"输入为空，生成空 workbook（Sheet: {sheet_name}）", verbose)
            wb.save(args.output)
        sys.exit(0)

    log(f"表头: {headers}", verbose)
    log(f"数据行数: {len(data_rows)}", verbose)

    # 执行排序
    try:
        sorted_indices = sort_data(headers, data_rows, rules, verbose)
    except Exception as e:
        # P1-4 修复：保留 traceback
        traceback.print_exc(file=sys.stderr)
        print(f"错误: 排序失败 - {e}", file=sys.stderr)
        sys.exit(1)

    log(f"排序完成，行顺序: {sorted_indices[:10]}{'...' if len(sorted_indices) > 10 else ''}", verbose)

    # 输出（P1-5：复用前面解析的 out_fmt，不再二次调用 detect_format）
    try:
        if out_fmt == "csv":
            write_csv(headers, data_rows, sorted_indices, args.output, verbose)
        else:
            write_xlsx(headers, data_rows, sorted_indices, meta, args.output, verbose)
    except Exception as e:
        # P1-4 修复：保留 traceback
        traceback.print_exc(file=sys.stderr)
        print(f"错误: 输出失败 - {e}", file=sys.stderr)
        sys.exit(1)

    print(f"✅ 排序完成: {args.output}")


if __name__ == "__main__":
    main()

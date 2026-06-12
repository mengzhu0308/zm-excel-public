"""Shared helpers for extract_template.py and write_back.py.

Centralizes:
  - select_worksheet(): pick a worksheet by name or fall back to active sheet
  - detect_header_row(): pick the row among the first 3 with the most non-empty cells
  - find_sample_row(): locate the last data row aligned to the header
  - copy_styles(): copy visual styles from one cell to another
  - load_workbook(): explicit data_only policy wrapper
  - normalize_header(): collapse internal whitespace, replace newlines, strip

Keeping these in one place avoids drift between extract_template.py and
write_back.py.
"""

from __future__ import annotations

import re
import sys
import zipfile
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

_WHITESPACE_RE = re.compile(r"\s+")


HEADER_SCAN_ROWS = 3

# File extensions this skill actually supports. openpyxl can only read/write
# .xlsx and .xlsm, so anything else should be rejected early with a clear
# message instead of letting openpyxl raise a confusing zipfile error.
SUPPORTED_EXCEL_EXTENSIONS = (".xlsx", ".xlsm")


def validate_excel_extension(path) -> None:
    """Raise ValueError if path does not have a supported Excel extension.

    Keeps the user-facing extension check in one place so extract_template.py
    and write_back.py stay aligned. The check is purely about the extension
    string — actual openpyxl errors are still surfaced as-is for malformed
    .xlsx files.
    """
    suffix = Path(path).suffix.lower()
    if suffix not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValueError(
            f"Unsupported file extension: '{suffix or '(none)'}'. "
            f"This skill only supports {', '.join(SUPPORTED_EXCEL_EXTENSIONS)}. "
            f"If you have a legacy .xls file, please convert it to .xlsx first."
        )


def select_worksheet(wb, sheet_name: Optional[str]):
    """Pick a worksheet by name, or fall back to the active sheet.

    When sheet_name is None and the workbook has multiple sheets, prints a
    stderr note listing the available sheets so the user can confirm the
    active sheet is the intended target. This keeps extract_template.py
    and write_back.py aligned on multi-sheet handling.

    Args:
        wb: An openpyxl Workbook.
        sheet_name: Optional worksheet name. If provided and exists, it is used.

    Returns:
        The selected openpyxl Worksheet.

    Raises:
        ValueError: If sheet_name is provided but not found, or if the workbook
                    has no active sheet.
    """
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        available = ", ".join(f"'{s}'" for s in wb.sheetnames)
        raise ValueError(
            f"Worksheet '{sheet_name}' not found. "
            f"Available sheets: {available}"
        )
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active worksheet")
    if len(wb.sheetnames) > 1:
        sheets_str = ", ".join(f"'{s}'" for s in wb.sheetnames)
        print(
            f"Note: File has multiple sheets ({sheets_str}). "
            f"Using active sheet '{ws.title}'. Use --sheet to specify another.",
            file=sys.stderr,
        )
    return ws


def detect_header_row(ws) -> int:
    """Auto-detect header row: pick the row among the first HEADER_SCAN_ROWS
    with the most non-empty cells.

    Raises:
        ValueError: If the worksheet has no readable rows.
    """
    def _non_empty_count(row_cells):
        return sum(1 for c in row_cells if c.value is not None)

    candidate_rows = []
    upper = min(HEADER_SCAN_ROWS, ws.max_row if ws.max_row else 0) + 1
    for r in range(1, max(2, upper)):
        candidate_rows.append((_non_empty_count(ws[r]), r))

    if not candidate_rows:
        raise ValueError("Excel file has no readable rows")

    _, header_row_idx = max(candidate_rows, key=lambda x: x[0])
    return header_row_idx


def find_sample_row(ws, header_cols: list[int]) -> Optional[int]:
    """Find the last row that looks like a data row (aligned to headers).

    A data row must have values in at least half of the columns listed
    in ``header_cols``. Callers compute ``header_cols`` from
    ``read_headers()`` once and pass the result in, so the headers list
    is not re-derived (and re-normalized) on every call. This both
    halves the work for large sheets and removes the implicit
    dependence on ``read_headers`` from inside this function.

    Args:
        ws: The openpyxl Worksheet.
        header_cols: 1-based column indices of the named header columns
                     (i.e. columns whose normalized header is non-empty).
                     Callers typically build this with
                     ``[c.column for c, name in zip(ws[row], headers) if name]``.

    Returns:
        The 1-based row index of the last data row, or None when no row
        meets the "at least half the named columns populated" rule.
    """
    if not header_cols:
        return None

    threshold = max(1, len(header_cols) // 2)
    for row_idx in range(ws.max_row, 0, -1):
        vals = sum(
            1
            for c in header_cols
            if ws.cell(row=row_idx, column=c).value is not None
        )
        if vals >= threshold:
            return row_idx
    return None


def copy_styles(src_cell, dst_cell) -> None:
    """Copy visual styles from src_cell to dst_cell.

    In addition to the per-cell style properties (font, border, fill,
    number_format, protection, alignment), the row height of the source
    cell's row is also copied to the destination cell's row. Without
    this, the new row gets openpyxl's default 15-point height even when
    the source row was sized to e.g. 30 points — a small but visible
    style break that SKILL.md and README.md promise "等样式" covers.
    """
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)
    # Row height lives on row_dimensions, not on the cell. Only set it
    # when the source row was explicitly sized; otherwise the openpyxl
    # default (None) would clobber whatever default the worksheet
    # already has.
    src_row_h = src_cell.parent.row_dimensions[src_cell.row].height
    if src_row_h is not None:
        dst_cell.parent.row_dimensions[dst_cell.row].height = src_row_h


def normalize_header(value) -> str:
    """Normalize a header cell value into a string for matching.

    - None -> ""
    - str -> collapse all runs of whitespace (including \\n) into a
      single space, then strip
    - other -> str(...)

    Collapsing internal whitespace ensures that headers like
    "姓 名" and "姓  名" are treated as the same field name when
    matching between Excel and Markdown.
    """
    if value is None:
        return ""
    return _WHITESPACE_RE.sub(" ", str(value)).strip()


def read_headers(ws, header_row_idx: int) -> list[str]:
    """Read the header row into a normalized list of strings.

    Empty strings are preserved in their original column position so that the
    column index in the headers list matches the column index in the sheet.
    """
    headers = []
    for cell in ws[header_row_idx]:
        headers.append(normalize_header(cell.value))
    return headers


def load_workbook(path: str, *, data_only: bool, keep_vba: bool = False):
    """Wrapper around openpyxl.load_workbook with explicit data_only and keep_vba policy.

    Use data_only=True when you need computed values (template extraction).
    Use data_only=False when you need formulas preserved (write-back).
    Pass keep_vba=True when writing back to a .xlsm that may contain VBA
    macros — without it, openpyxl strips the vbaProject.bin on save
    (silent data loss; SKILL.md and README.md promise .xlsm support
    includes macro-bearing files).
    """
    return openpyxl.load_workbook(path, data_only=data_only, keep_vba=keep_vba)


def assert_headers_present(headers: list[str]) -> None:
    """Raise ValueError when no named headers are present.

    A single helper to keep extract_template.py and write_back.py aligned:
    both scripts must refuse to continue when the detected header row is
    entirely empty or whitespace-only (which normalize_header collapses
    to ""). Without this guard, write_back would silently produce an
    all-None new row.
    """
    if not headers or all(h == "" for h in headers):
        raise ValueError("No headers found")


# Sentinel substring from raise_auto_increment_limit's 9999-limit
# FileExistsError. When this substring is present in a FileExistsError
# message, localize_error renders the user-facing Chinese text that
# SKILL.md and README.md promise; otherwise it falls through to the
# raw English message so the user can still see the original context.
# Single source of truth: this module defines the marker AND is the
# only place that raises the marker-bearing FileExistsError, so a
# future wording change in either place stays in lockstep.
_AUTO_INCREMENT_LIMIT_MARKER = "Auto-increment reached the 9999 limit"

# Cap for output auto-increment (<stem>_增加一行.xlsx → _2.xlsx → ... → _9999.xlsx).
# Mirrors the limit documented in SKILL.md and README.md. Lives here so
# the marker, the cap, and the raise helper all live in one module.
MAX_AUTO_INCREMENT = 9999


def raise_auto_increment_limit(excel_file, base, total_count=None) -> None:
    """Raise the 9999-limit FileExistsError from a single source of truth.

    Constructed once in _common so that the marker substring, the limit
    value, and the wording of the FileExistsError all stay in lockstep
    with localize_error. write_back._resolve_output_path used to build
    the same sentence inline; centralizing it removes the cross-file
    hidden contract (B-P1-3).

    Args:
        excel_file: The source Path; only ``.stem`` / ``.suffix`` and
                    ``.parent`` are read to build the diagnostic.
        base: The base output Path that already exists and triggered
              the auto-increment.
        total_count: Optional pre-computed count of allocated indices
                     to avoid the second glob in the hot error path.
                     When None, computed via ``excel_file.parent.glob``.

    Raises:
        FileExistsError: Always; the message embeds
                         ``_AUTO_INCREMENT_LIMIT_MARKER`` so that
                         ``localize_error`` renders the Chinese
                         wording the SKILL.md and README.md promise.
    """
    existing_indices: list[int] = []
    for p in excel_file.parent.glob(
        f"{excel_file.stem}_增加一行_*{excel_file.suffix}"
    ):
        tail = p.stem.rsplit("_", 1)[-1]
        if tail.isdigit():
            existing_indices.append(int(tail))
    existing_indices.sort()
    sample = existing_indices[:20]
    extra = len(existing_indices) - len(sample)
    indices_repr = ", ".join(str(i) for i in sample)
    if extra > 0:
        indices_repr += f", ... (+{extra} more)"
    if total_count is None:
        total_count = len(existing_indices)
    raise FileExistsError(
        f"Auto-increment reached the {MAX_AUTO_INCREMENT} limit for outputs matching "
        f"{base.name}*. No available filename of the form "
        f"'{excel_file.stem}_增加一行_<n>.xlsx' was found. "
        f"Already-allocated indices: {indices_repr} "
        f"(total: {total_count}). "
        f"Clean up the target directory or pass --output explicitly "
        f"to write to a different path."
    )


def localize_error(exc: Exception) -> str:
    """Render an exception as the user-facing message.

    Centralizes the FileExistsError (9999 auto-increment limit) and
    PermissionError (Excel locked) Chinese wording that SKILL.md and
    README.md promise. The three CLI entrypoints (add_one_row.py,
    extract_template.py main(), write_back.py main()) all funnel
    exceptions through this helper so the user-facing contract is
    defined in exactly one place.

    Args:
        exc: The exception to localize.

    Returns:
        The user-facing message string (no trailing newline).
    """
    if isinstance(exc, FileExistsError):
        msg = str(exc)
        if _AUTO_INCREMENT_LIMIT_MARKER in msg:
            return (
                "已达自动递增上限 9999，请清理目标目录或改用 --output 显式指定。"
            )
        return f"Error: {msg}"
    if isinstance(exc, PermissionError):
        fname = exc.filename or "<unknown>"
        return f"Error: Excel 文件被其他程序占用或无写权限: {fname}。请关闭 Excel 后重试。"
    if isinstance(exc, FileNotFoundError):
        return f"Error: {exc}"
    # Corrupted .xlsx/.xlsm or wrong-format files: openpyxl raises
    # InvalidFileException, while low-level zipfile.BadZipFile leaks through
    # when openpyxl can't even open the archive. Map both to a single
    # Chinese message so the user sees the same root cause regardless of
    # which layer the error escapes from.
    if isinstance(exc, zipfile.BadZipFile):
        fname = getattr(exc, "filename", None) or "<unknown>"
        return (
            f"Error: 无法解析 Excel 文件: {fname}。"
            f"文件可能损坏、不是有效的 .xlsx/.xlsm，或被其他程序占用。"
            f"如需处理 .xls 请先用其他工具转换为 .xlsx。"
        )
    if isinstance(exc, InvalidFileException):
        fname = getattr(exc, "filename", None) or "<unknown>"
        return (
            f"Error: openpyxl 无法识别该文件: {fname}。"
            f"请确认它是有效的 .xlsx 或 .xlsm 文件。"
        )
    return f"Error: {exc}"


def _setup_path() -> None:
    """Insert the scripts/ directory at the front of sys.path.

    Called by every CLI entrypoint before importing sibling modules
    (`_common`, `extract_template`, `write_back`). Without this, scripts
    only resolve each other when launched as a top-level script (where
    Python sets sys.path[0] to the script's directory); running via
    `python3 -m scripts.add_one_row` or with a stripped PYTHONPATH
    surfaces ImportError for the sibling modules.
    """
    sys.path.insert(0, str(Path(__file__).parent))


__all__ = [
    "HEADER_SCAN_ROWS",
    "SUPPORTED_EXCEL_EXTENSIONS",
    "MAX_AUTO_INCREMENT",
    "select_worksheet",
    "detect_header_row",
    "find_sample_row",
    "copy_styles",
    "normalize_header",
    "read_headers",
    "load_workbook",
    "validate_excel_extension",
    "assert_headers_present",
    "localize_error",
    "raise_auto_increment_limit",
    "_setup_path",
]

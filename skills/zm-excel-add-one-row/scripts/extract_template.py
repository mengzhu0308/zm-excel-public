#!/usr/bin/env python3
"""Extract Excel headers and sample row into a Markdown template."""

import sys
from pathlib import Path
from typing import Optional

from _common import (
    _setup_path,
    assert_headers_present,
    detect_header_row,
    find_sample_row,
    load_workbook,
    localize_error,
    read_headers,
    select_worksheet,
    validate_excel_extension,
)

# Make sibling imports resolvable when launched as a top-level script
# or via `python3 -m scripts.extract_template`.
_setup_path()

try:
    import openpyxl  # noqa: F401  (imported for error messaging in __main__)
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def extract_template(
    excel_path: str,
    output_md: Optional[str] = None,
    sheet_name: Optional[str] = None,
    *,
    force: bool = False,
) -> str:
    """Extract headers and last row sample from Excel into Markdown template.

    Args:
        excel_path: Path to the Excel file.
        output_md: Optional explicit output path for the Markdown file.
                   Defaults to <excel_stem>_row_template.md in the same directory.
        sheet_name: Optional worksheet name to extract from.
                    Defaults to the active worksheet.
        force: If False (default), refuse to overwrite an existing template.
               If True, overwrite the existing file.

    Returns:
        Path to the generated Markdown file.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        FileExistsError: If the output Markdown file already exists and force is False.
        PermissionError: If the Excel file is locked or unwritable.
        ValueError: If the worksheet cannot be selected or no headers are found.
    """
    excel_file = Path(excel_path)
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    # Reject unsupported extensions early: .xls will produce confusing
    # zipfile errors if we let openpyxl try to open it.
    validate_excel_extension(excel_file)

    # openpyxl Workbook does NOT implement the context manager protocol,
    # so we release the file handle explicitly in a try/finally to avoid
    # file-lock leaks (notably on Windows).
    wb = load_workbook(str(excel_file), data_only=True)
    try:
        ws = select_worksheet(wb, sheet_name)

        header_row_idx = detect_header_row(ws)
        headers = read_headers(ws, header_row_idx)

        assert_headers_present(headers)

        header_cols = [c.column for c, name in zip(ws[header_row_idx], headers) if name != ""]
        sample_row_idx = find_sample_row(ws, header_cols)
        sample_row: dict[int, str] = {}
        if sample_row_idx is not None:
            for idx, cell in enumerate(ws[sample_row_idx], start=1):
                val = cell.value
                if val is not None:
                    sample_row[idx] = str(val)

        # Determine output path
        if output_md:
            md_path = Path(output_md)
        else:
            md_path = excel_file.with_name(f"{excel_file.stem}_row_template.md")

        # Refuse to overwrite an existing template unless force=True.
        # This protects user-edited content during cross-session resume.
        if md_path.exists() and not force:
            raise FileExistsError(
                f"Template already exists: {md_path}. "
                f"Refusing to overwrite. Use --force to overwrite explicitly."
            )

        lines: list[str] = []
        lines.append("## Excel 数据录入模板")
        lines.append("")
        lines.append(f"> 来源文件: `{excel_file.name}`")
        lines.append(f"> 工作表: `{ws.title}`")
        lines.append("")
        lines.append("请在下方的每个字段下填写实际内容，将每段末尾的占位符替换为你的数据。")
        lines.append("示例值来自表格中的已有数据，供你参考格式。")
        lines.append("")
        lines.append("---")
        lines.append("")

        for idx, header in enumerate(headers, start=1):
            if header == "":
                continue
            lines.append(f"# {header}")
            if idx in sample_row:
                lines.append(f"<!-- 示例: {sample_row[idx]} -->")
            lines.append("[在此填写]")
            lines.append("")

        md_path.write_text("\n".join(lines), encoding="utf-8")
    finally:
        wb.close()
    return str(md_path)

if __name__ == "__main__":
    print(
        "deprecated: 直接运行 `extract_template.py` 已废弃，请改用 `add_one_row.py extract`；"
        "历史 argparse 入口与 main() 已从该脚本移除，extract 逻辑以 add_one_row.py 为唯一入口",
        file=sys.stderr,
    )
    sys.exit(1)

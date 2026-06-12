#!/usr/bin/env python3
"""Parse filled Markdown template and append a new row to Excel."""

import re
import sys
from pathlib import Path
from typing import Optional

from _common import (
    _setup_path,
    MAX_AUTO_INCREMENT,
    assert_headers_present,
    copy_styles,
    detect_header_row,
    find_sample_row,
    load_workbook,
    localize_error,
    raise_auto_increment_limit,
    read_headers,
    select_worksheet,
    validate_excel_extension,
)

# Make sibling imports resolvable when launched as a top-level script
# or via `python3 -m scripts.write_back`.
_setup_path()

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


PLACEHOLDER = "[在此填写]"

# Strings with more than this many digits are likely IDs/phone numbers/account
# numbers and should bypass the int() coercion. The threshold is set to 10,
# deliberately below Excel's 15-digit precision ceiling: a 10–15 digit band
# leaves enough headroom that the 15-digit float64 round-trip can't quietly
# flip the trailing digits while still catching the common cases
# (身份证 18 位, 手机号 11 位, 银行卡 16–19 位, 工号 > 10 位). This is the
# single source of truth for that policy; SKILL.md 注意事项 / README.md Q&A
# mirror it but do not redefine it.
LONG_DIGIT_THRESHOLD = 10

# Cap for output auto-increment is re-exported from `_common` at the top
# of this module; see `_common.MAX_AUTO_INCREMENT` and
# `_common.raise_auto_increment_limit` for the single source of truth.


def _safe_int(text: str) -> Optional[object]:
    """Parse text as int; return None on failure."""
    try:
        return int(text)
    except ValueError:
        return None


def _safe_float(text: str) -> Optional[object]:
    """Parse text as float; return None on failure."""
    try:
        return float(text)
    except ValueError:
        return None


def _sample_value_type(sample: Optional[str]) -> str:
    """Classify a sample value as 'int', 'float', or 'string'.

    Used to decide whether a user-entered value should be coerced into a
    numeric type. If the sample is None or unparseable, default to 'string'
    to avoid silently corrupting data.
    """
    if sample is None:
        return "string"
    if _safe_int(sample) is not None:
        return "int"
    if _safe_float(sample) is not None:
        return "float"
    return "string"


# Sample-type → numeric parser. Looks up the parser for the sample row's
# detected type and runs it on the user-entered value. Keeps the conversion
# rules declarative so future "decimal" / "currency" support only needs a
# new entry here rather than a new branch in _convert_value.
_TYPE_PARSERS = {
    "int": _safe_int,
    "float": _safe_float,
    "string": lambda _text: None,  # never coerce; caller treats None as "keep as str"
}


def _convert_value(value: str, sample: Optional[str] = None) -> object:
    """Convert a user-entered string into an int / float / str for Excel.

    Rules (conservative — preserve data fidelity over cleverness):
      - empty -> None (cell stays blank)
      - long pure-digit strings (>10 digits) -> str (avoid 15-digit loss,
        preserve leading zeros, avoid scientific notation surprises)
      - pure-digit strings that start with '0' (e.g. '007', '01') -> str
        (Excel would otherwise drop the leading zero)
      - strings containing non-numeric characters -> str
      - if a sample was provided, only convert to int/float when the sample
        is of the matching numeric type; otherwise keep as str
      - otherwise try int, then float; if both fail, keep as str

    The "coerce by sample type" path is driven by _TYPE_PARSERS so adding
    a new numeric type is a one-line change in that table rather than a
    new branch here.
    """
    if value == "":
        return None

    # Strip whitespace for inspection but do not mutate the input yet.
    stripped = value.strip()
    if stripped == "":
        return None

    sample_type = _sample_value_type(sample)

    # Sign-stripped digit check: Python 3's str.isdigit() returns False for
    # values with a leading + or - (e.g. "-5".isdigit() == False), so we
    # lstrip the sign characters before testing digit-only-ness.
    sign_stripped = stripped.lstrip("+-")
    is_pure_digits = sign_stripped.isdigit()

    if is_pure_digits:
        # Long pure-digit guard: avoid Excel 15-digit precision loss and
        # preserve any leading zeros that exceed 10 digits.
        if len(sign_stripped) > LONG_DIGIT_THRESHOLD:
            return value
        # Leading-zero guard: "007" passes isdigit, but Excel drops the
        # leading zero if we coerce to int. We must check the original
        # digit string (after sign strip) explicitly.
        if sign_stripped.startswith("0") and len(sign_stripped) > 1:
            return value
        parser = _TYPE_PARSERS.get(sample_type)
        if parser is not None:
            parsed = parser(stripped)
            if parsed is not None:
                return parsed
        # Sample type is "string" or parser failed — keep the original
        # text intact.
        return value

    # Non-pure-digit input: only coerce to float when the sample is float.
    if sample_type == "float":
        parsed = _safe_float(stripped)
        if parsed is not None:
            return parsed
    return value


_PLACEHOLDER_PREFIX_RE = re.compile(r"^\s*" + re.escape(PLACEHOLDER) + r"\s*")
_HEADING_RE = re.compile(r"^#\s+(.+?)\n(.*?)(?=\n#\s|\Z)", re.MULTILINE | re.DOTALL)
_SHEET_RE = re.compile(r"工作表:\s*`([^`]+)`")
_COMMENT_RE = re.compile(r"<!--.*?-->", re.DOTALL)


def parse_markdown(md_path: str) -> tuple[dict[str, str], Optional[str]]:
    """Parse a Markdown template and extract field values and sheet name.

    Returns a tuple of (field_values, sheet_name) where:
    - field_values: dict mapping header name to user-provided value
    - sheet_name: sheet name from Markdown template (or None if not found)

    Empty or placeholder values are returned as empty strings. A line that
    equals the placeholder (or starts with the placeholder) is treated as
    "未填写" — the leading placeholder prefix is stripped and the remaining
    text is used as the actual value. Only the first non-empty line under
    each heading is used; extra content (tables, etc.) is ignored.
    """
    content = Path(md_path).read_text(encoding="utf-8")

    sheet_name: Optional[str] = None
    sheet_match = _SHEET_RE.search(content)
    if sheet_match:
        sheet_name = sheet_match.group(1).strip()

    result: dict[str, str] = {}
    for match in _HEADING_RE.finditer(content):
        header = match.group(1).strip()
        body = match.group(2).strip()
        # Remove HTML comments (example hints)
        body = _COMMENT_RE.sub("", body).strip()
        # Only take the first non-empty line; ignore tables or extra lines below
        lines = [line.strip() for line in body.splitlines() if line.strip()]
        if not lines:
            result[header] = ""
            continue
        first = lines[0]
        if first == PLACEHOLDER:
            result[header] = ""
            continue
        # Strip a leading "[在此填写]" prefix that the user forgot to remove.
        rest = _PLACEHOLDER_PREFIX_RE.sub("", first)
        if rest == "":
            result[header] = ""
        else:
            result[header] = rest

    return result, sheet_name


def _print_preview(
    headers: list[str],
    new_row: list[object],
    md_only_fields: list[str],
    unmatched: list[str],
    next_row: int,
) -> None:
    """Print a formatted preview of the row to be appended.

    Surfaces the row number the new row will land on FIRST (above the
    field-by-field table) so users running --dry-run can verify the
    insertion point immediately (e.g. when the sheet contains hidden
    rows or filters that change max_row). The field list is then
    numbered with a 1-based index for easy cross-referencing.
    """
    print("=" * 50)
    print(f"预览：即将追加到 Excel 第 {next_row} 行的数据")
    print("=" * 50)

    max_header_len = max((len(h) for h in headers), default=0)
    for idx, (header, value) in enumerate(zip(headers, new_row), start=1):
        display = "(空)" if value is None else str(value)
        print(
            f"  {str(idx).rjust(2)}. {header.ljust(max_header_len)}  →  {display}"
        )

    print("=" * 50)

    if md_only_fields:
        print(f"警告: Markdown 中有但 Excel 表头没有的字段（已忽略）: {md_only_fields}")
    if unmatched:
        print(f"警告: Excel 表头中有但 Markdown 中没有的字段（留空）: {unmatched}")


def _resolve_output_path(
    excel_file: Path,
    explicit_output: Optional[str] = None,
    force: bool = False,
) -> Path:
    """Determine the output xlsx path, preserving prior appended rows.

    Priority:
      1. explicit_output (--output CLI / function arg) — used as-is;
         if it already exists and not force, raise FileExistsError.
         Note: --force is ONLY honored in this branch; the auto-increment
         branch below intentionally never overwrites an existing
         <stem>_增加一行*.xlsx, so passing --force without --output will
         still increment rather than clobber the first slot.
      2. <stem>_增加一行.xlsx — if exists, increment to
         <stem>_增加一行_2.xlsx, <stem>_增加一行_3.xlsx, ... so that
         consecutive calls on the same source do NOT clobber the
         previously generated output file. The scan starts at _2 and
         returns the first gap; it does NOT fill intermediate holes
         left by manually deleted files (by design — see SKILL.md).
    """
    if explicit_output:
        out = Path(explicit_output)
        if out.exists() and not force:
            raise FileExistsError(
                f"Output file already exists: {out}. "
                f"Refusing to overwrite. Use --force to overwrite explicitly."
            )
        return out

    base = excel_file.with_name(f"{excel_file.stem}_增加一行{excel_file.suffix}")
    if not base.exists():
        return base
    n = 2
    while True:
        candidate = excel_file.with_name(
            f"{excel_file.stem}_增加一行_{n}{excel_file.suffix}"
        )
        if not candidate.exists():
            return candidate
        n += 1
        if n > MAX_AUTO_INCREMENT:
            # Defer the FileExistsError construction (and the existing-index
            # glob) to _common.raise_auto_increment_limit so the marker
            # substring, the cap, and the wording all live in one place.
            # See _common._AUTO_INCREMENT_LIMIT_MARKER for the marker.
            raise_auto_increment_limit(excel_file, base)


def write_back(
    excel_path: str,
    md_path: str,
    *,
    dry_run: bool = False,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
    force: bool = False,
) -> Optional[int]:
    """Append data from Markdown template to a new Excel file.

    Args:
        excel_path: Path to the original Excel file.
        md_path: Path to the filled Markdown template.
        dry_run: If True, only print preview without writing any file.
        sheet_name: Optional worksheet name. If provided, overrides the sheet
                    recorded in the Markdown template.
        output_path: Optional explicit output xlsx path. If omitted, the
                     script writes to <stem>_增加一行.xlsx; if that file
                     already exists, it auto-increments to
                     <stem>_增加一行_2.xlsx, _3.xlsx, etc.
        force: If True, allow overwriting an existing output file when
               output_path is explicit. Ignored when output_path is None
               (auto-increment never overwrites).

    Returns:
        The row number of the newly appended row, or None if dry_run.
    """
    excel_file = Path(excel_path)
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    # Reject unsupported extensions early: .xls will produce confusing
    # zipfile errors if we let openpyxl try to open it.
    validate_excel_extension(excel_file)
    md_file = Path(md_path)
    if not md_file.exists():
        raise FileNotFoundError(f"Markdown file not found: {md_path}")

    field_values, md_sheet_name = parse_markdown(md_path)
    if not field_values:
        raise ValueError(
            "No fields found in the Markdown file. "
            "Make sure it uses `# FieldName` format."
        )

    # Priority: CLI --sheet > Markdown template > active sheet
    effective_sheet = sheet_name if sheet_name is not None else md_sheet_name

    # Load workbook once and reuse for preview, append, and save.
    # openpyxl Workbook does NOT implement the context manager protocol;
    # release the file handle explicitly in try/finally to avoid file-lock
    # leaks (notably on Windows).
    wb = load_workbook(str(excel_file), data_only=False, keep_vba=True)
    try:
        ws = select_worksheet(wb, effective_sheet)

        header_row_idx = detect_header_row(ws)
        headers = read_headers(ws, header_row_idx)
        assert_headers_present(headers)

        # Build a parallel sample-row lookup so we can preserve numeric
        # coercion intent. sample_by_header maps Excel header -> sample
        # value (str).
        header_cols = [c.column for c, name in zip(ws[header_row_idx], headers) if name != ""]
        sample_row_idx = find_sample_row(ws, header_cols)
        sample_by_header: dict[str, str] = {}
        if sample_row_idx is not None:
            for cell in ws[sample_row_idx]:
                if cell.value is None:
                    continue
                # Reuse the already-fetched headers list rather than
                # re-calling read_headers here; the latter would do
                # redundant normalization work and could drift from the
                # headers used by the rest of this function.
                header_name = headers[cell.column - 1] \
                    if cell.column - 1 < len(headers) else ""
                if header_name:
                    sample_by_header[header_name] = str(cell.value)

        new_row: list[object] = []
        unmatched: list[str] = []
        for header in headers:
            if header == "":
                new_row.append(None)
                continue
            if header in field_values:
                new_row.append(_convert_value(field_values[header], sample_by_header.get(header)))
            else:
                unmatched.append(header)
                new_row.append(None)

        md_only_fields = sorted(set(field_values.keys()) - set(headers))

        # Determine insertion point: immediately after the last data row,
        # not at the worksheet tail (which may contain blank rows or footer).
        if sample_row_idx is not None:
            next_row = sample_row_idx + 1
        else:
            next_row = header_row_idx + 1

        # Inform user when CLI overrides Markdown sheet
        if sheet_name is not None and md_sheet_name is not None and sheet_name != md_sheet_name:
            print(f"提示: 命令行指定工作表 '{sheet_name}' 覆盖了模板中的 '{md_sheet_name}'")

        _print_preview(headers, new_row, md_only_fields, unmatched, next_row)

        # Surface the "no sample row → no style copy" warning before the
        # dry-run return so users running --dry-run also see it.
        if sample_row_idx is None:
            print(
                "提示: 源表只有表头没有数据行，新行不会复制任何样式（保持默认）。",
                file=sys.stderr,
            )

        if dry_run:
            print("\n[dry-run] 未执行写入。如需写入，请去掉 --dry-run 参数重新执行。")
            return None

        try:
            output_file = _resolve_output_path(excel_file, output_path, force)
        except FileExistsError as e:
            print(f"Error: {e}")
            raise

        # Write cell-by-cell so the new row lands right after the last data
        # row instead of being pushed to the worksheet tail.
        for col_idx, value in enumerate(new_row, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)
        new_row_num = next_row

        if sample_row_idx is not None:
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=sample_row_idx, column=col_idx)
                dst_cell = ws.cell(row=new_row_num, column=col_idx)
                copy_styles(src_cell, dst_cell)

        wb.save(str(output_file))
        print(f"\n新表格已保存: {output_file}")
        print(f"数据追加到第 {new_row_num} 行")
        return new_row_num
    finally:
        wb.close()

if __name__ == "__main__":
    print(
        "deprecated: 直接运行 `write_back.py` 已废弃，请改用 `add_one_row.py write`；"
        "历史 argparse 入口与 main() 已从该脚本移除，write 逻辑以 add_one_row.py 为唯一入口",
        file=sys.stderr,
    )
    sys.exit(1)
